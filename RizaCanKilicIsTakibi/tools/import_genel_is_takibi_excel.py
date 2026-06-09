from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
import unicodedata
import uuid
from pathlib import Path

from openpyxl import load_workbook

TASK_BOARD_GENEL = 0
TASK_BOARD_ACIL = 1

ACTION_CATEGORY_AKSIYON = 0
ACTION_CATEGORY_AKSIYONA_EKLENECEKLER = 1

MISSING_MEDIUM_DIJITAL = 0
MISSING_MEDIUM_FIZIKI = 1
MISSING_MEDIUM_FIZIKI_VE_DIJITAL = 2


def now_iso() -> str:
    return dt.datetime.now().isoformat()


def normalize_text(value: str | None) -> str:
    if value is None:
        return ""

    text = value.strip().lower().replace("ı", "i").replace("İ", "i")
    decomposed = unicodedata.normalize("NFD", text)
    stripped = "".join(ch for ch in decomposed if unicodedata.category(ch) != "Mn")
    return " ".join(stripped.split())


def text_or_empty(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_missing_medium(value: str) -> tuple[int, str]:
    normalized = normalize_text(value)
    has_dijital = "dijital" in normalized
    has_fiziki = "fiziki" in normalized or "fiziksel" in normalized or "fizik" in normalized

    if has_dijital and has_fiziki:
        return MISSING_MEDIUM_FIZIKI_VE_DIJITAL, "Fiziksel + Dijital"
    if has_dijital:
        return MISSING_MEDIUM_DIJITAL, "Dijital"
    if has_fiziki:
        return MISSING_MEDIUM_FIZIKI, "Fiziksel"

    return MISSING_MEDIUM_FIZIKI, "Fiziksel"


def find_sheet(workbook, sheet_name: str):
    target = normalize_text(sheet_name)
    for ws in workbook.worksheets:
        if normalize_text(ws.title) == target:
            return ws

    available = ", ".join(ws.title for ws in workbook.worksheets)
    raise ValueError(f"'{sheet_name}' sayfası bulunamadı. Mevcut sayfalar: {available}")


def read_genel_tasks(ws) -> list[tuple[int, str, str]]:
    rows: list[tuple[int, str, str]] = []
    current_board: int | None = None

    for row_index in range(1, ws.max_row + 1):
        raw_title = text_or_empty(ws.cell(row=row_index, column=1).value)
        raw_description = text_or_empty(ws.cell(row=row_index, column=2).value)
        normalized_title = normalize_text(raw_title)

        if normalized_title == "acil yapilacak is":
            current_board = TASK_BOARD_ACIL
            continue
        if normalized_title == "genel yapilacak is":
            current_board = TASK_BOARD_GENEL
            continue
        if not raw_title:
            continue
        if current_board is None:
            continue

        rows.append((current_board, raw_title, raw_description))

    return rows


def read_action_entries(ws) -> list[tuple[int, str, str, str]]:
    rows: list[tuple[int, str, str, str]] = []
    current_district = ""

    for row_index in range(2, ws.max_row + 1):
        district = text_or_empty(ws.cell(row=row_index, column=1).value)
        owner_parcel = text_or_empty(ws.cell(row=row_index, column=2).value)
        work_text = text_or_empty(ws.cell(row=row_index, column=3).value)

        if district:
            current_district = district
        if not owner_parcel and not work_text:
            continue

        rows.append((ACTION_CATEGORY_AKSIYON, current_district, owner_parcel, work_text))

    return rows


def read_action_to_add_entries(ws, district: str) -> list[tuple[int, str, str, str]]:
    rows: list[tuple[int, str, str, str]] = []

    for row_index in range(2, ws.max_row + 1):
        owner_parcel = text_or_empty(ws.cell(row=row_index, column=1).value)
        work_text = text_or_empty(ws.cell(row=row_index, column=2).value)
        if not owner_parcel and not work_text:
            continue

        rows.append((ACTION_CATEGORY_AKSIYONA_EKLENECEKLER, district, owner_parcel, work_text))

    return rows


def read_missing_project_entries(ws) -> list[tuple[str, str, int, str, str, str]]:
    rows: list[tuple[str, str, int, str, str, str]] = []

    for row_index in range(2, ws.max_row + 1):
        ada_parsel = text_or_empty(ws.cell(row=row_index, column=1).value)
        yapi_sahibi = text_or_empty(ws.cell(row=row_index, column=2).value)
        medium_text_raw = text_or_empty(ws.cell(row=row_index, column=3).value)
        missing_project_text = text_or_empty(ws.cell(row=row_index, column=4).value)
        description = text_or_empty(ws.cell(row=row_index, column=5).value)

        if not any((ada_parsel, yapi_sahibi, medium_text_raw, missing_project_text, description)):
            continue

        medium, medium_label = parse_missing_medium(medium_text_raw)
        rows.append((ada_parsel, yapi_sahibi, medium, medium_label, missing_project_text, description))

    return rows


def backup_database(db_path: Path) -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = db_path.with_name(f"{db_path.stem}.before_excel_import_{timestamp}{db_path.suffix}")

    source = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    destination = sqlite3.connect(str(backup_path))
    try:
        source.backup(destination)
    finally:
        destination.close()
        source.close()

    return backup_path


def get_next_task_sort_order(cursor: sqlite3.Cursor, board_type: int) -> int:
    cursor.execute("SELECT COALESCE(MAX(SortOrder), -1) FROM Tasks WHERE BoardType = ?", (board_type,))
    return int(cursor.fetchone()[0]) + 1


def get_next_action_display_order(cursor: sqlite3.Cursor, category: int, district: str) -> int:
    cursor.execute(
        "SELECT COALESCE(MAX(DisplayOrder), -1) FROM ActionEntries WHERE Category = ? AND District = ?",
        (category, district),
    )
    return int(cursor.fetchone()[0]) + 1


def get_next_missing_display_order(cursor: sqlite3.Cursor) -> int:
    cursor.execute("SELECT COALESCE(MAX(DisplayOrder), -1) FROM MissingProjectEntries")
    return int(cursor.fetchone()[0]) + 1


def import_workbook(workbook_path: Path, db_path: Path, aksiyona_eklenecekler_district: str) -> dict[str, int | str]:
    wb = load_workbook(workbook_path, data_only=True)

    ws_genel = find_sheet(wb, "GENEL İŞLER")
    ws_aksiyon = find_sheet(wb, "AKSİYON")
    ws_aksiyona_eklenecekler = find_sheet(wb, "AKSİYONA EKLENECEKLER")
    ws_eksik_proje = find_sheet(wb, "EKSİK PROJE")

    genel_rows = read_genel_tasks(ws_genel)
    aksiyon_rows = read_action_entries(ws_aksiyon)
    aksiyona_eklenecekler_rows = read_action_to_add_entries(ws_aksiyona_eklenecekler, aksiyona_eklenecekler_district)
    eksik_proje_rows = read_missing_project_entries(ws_eksik_proje)

    backup_path = backup_database(db_path)

    connection = sqlite3.connect(str(db_path), timeout=30)
    try:
        with connection:
            cursor = connection.cursor()

            next_sort_order = {
                TASK_BOARD_ACIL: get_next_task_sort_order(cursor, TASK_BOARD_ACIL),
                TASK_BOARD_GENEL: get_next_task_sort_order(cursor, TASK_BOARD_GENEL),
            }

            inserted_urgent = 0
            inserted_general = 0
            for board_type, title, description in genel_rows:
                created_at = now_iso()
                sort_order = next_sort_order[board_type]
                next_sort_order[board_type] += 1

                cursor.execute(
                    """
                    INSERT INTO Tasks (Id, Title, Description, DueDate, CreatedAt, UpdatedAt, BoardType, SortOrder)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid.uuid4()),
                        title,
                        description,
                        None,
                        created_at,
                        created_at,
                        board_type,
                        sort_order,
                    ),
                )

                if board_type == TASK_BOARD_ACIL:
                    inserted_urgent += 1
                else:
                    inserted_general += 1

            action_display_order_cache: dict[tuple[int, str], int] = {}
            inserted_action = 0
            inserted_to_add = 0
            for category, district, owner_parcel, work_text in aksiyon_rows + aksiyona_eklenecekler_rows:
                key = (category, district)
                if key not in action_display_order_cache:
                    action_display_order_cache[key] = get_next_action_display_order(cursor, category, district)

                display_order = action_display_order_cache[key]
                action_display_order_cache[key] += 1
                created_at = now_iso()

                cursor.execute(
                    """
                    INSERT INTO ActionEntries (Id, Category, District, OwnerParcelText, WorkText, DisplayOrder, CreatedAt, UpdatedAt)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid.uuid4()),
                        category,
                        district,
                        owner_parcel,
                        work_text,
                        display_order,
                        created_at,
                        created_at,
                    ),
                )

                if category == ACTION_CATEGORY_AKSIYON:
                    inserted_action += 1
                else:
                    inserted_to_add += 1

            missing_display_order = get_next_missing_display_order(cursor)
            inserted_missing = 0
            for ada_parsel, yapi_sahibi, medium, medium_label, missing_project_text, description in eksik_proje_rows:
                created_at = now_iso()
                cursor.execute(
                    """
                    INSERT INTO MissingProjectEntries
                    (Id, AdaParsel, YapiSahibi, RecordMedium, RecordMediumText, MissingProjectText, Description, DisplayOrder, CreatedAt, UpdatedAt)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid.uuid4()),
                        ada_parsel,
                        yapi_sahibi,
                        medium,
                        medium_label,
                        missing_project_text,
                        description,
                        missing_display_order,
                        created_at,
                        created_at,
                    ),
                )
                missing_display_order += 1
                inserted_missing += 1

        return {
            "backup_path": str(backup_path),
            "inserted_urgent": inserted_urgent,
            "inserted_general": inserted_general,
            "inserted_action": inserted_action,
            "inserted_to_add": inserted_to_add,
            "inserted_missing": inserted_missing,
        }
    finally:
        connection.close()


def resolve_default_db_path() -> Path:
    local_app_data = Path.home() / "AppData" / "Local"
    return local_app_data / "RizaCanKilicIsTakibi" / "tasks.db"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="GENEL İŞ TAKİBİ.xlsx verilerini uygulamanın SQLite veritabanına ekler.")
    parser.add_argument(
        "--excel",
        default=r"C:/Users/rizac/Masaüstü/İŞ TAKİBİ ÖRNEK EXCELLER/GENEL İŞ TAKİBİ.xlsx",
        help="Excel dosya yolu",
    )
    parser.add_argument(
        "--db",
        default=str(resolve_default_db_path()),
        help="tasks.db dosya yolu",
    )
    parser.add_argument(
        "--aksiyona-eklenecekler-ilce",
        default="MERKEZ",
        help="AKSİYONA EKLENECEKLER sayfası için ilçe değeri",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel)
    db_path = Path(args.db)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel bulunamadı: {excel_path}")
    if not db_path.exists():
        raise FileNotFoundError(f"Veritabanı bulunamadı: {db_path}")

    result = import_workbook(
        workbook_path=excel_path,
        db_path=db_path,
        aksiyona_eklenecekler_district=args.aksiyona_eklenecekler_ilce.strip() or "MERKEZ",
    )

    print("Aktarım tamamlandı.")
    print(f"Yedek: {result['backup_path']}")
    print(f"Genel İş Takibi / ACİL eklenen: {result['inserted_urgent']}")
    print(f"Genel İş Takibi / GENEL eklenen: {result['inserted_general']}")
    print(f"Aksiyon eklenen: {result['inserted_action']}")
    print(f"Aksiyona Eklenecekler eklenen: {result['inserted_to_add']}")
    print(f"Eksik Proje eklenen: {result['inserted_missing']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
